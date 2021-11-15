import openpyxl
from math import sin

wb = openpyxl.Workbook()

tab = wb.active
tab.title = "MeineErsteTabelle"

tab['A1'] = 'x'
tab['B1'] = 'f(x) = x * sin(x)'
row = 2

for i in range(-31, 32):
    x = i/10.0
    tab.cell(row, 1).value = x
    tab.cell(row, 2).value = x * sin(x)
    row += 1

wb.save("Datei17.xlsx")

